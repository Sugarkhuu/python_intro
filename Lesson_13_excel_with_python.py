#This time - openpyxl, xlrd, xlwt, XlsxWriter, csv

#One day - PyXLL, xlwings, pywin32 - like VBA in Excel

# Source
#http://zetcode.com/articles/openpyxl/
#https://openpyxl.readthedocs.io/en/stable/usage.html
#Further: https://www.pyxll.com/blog/tools-for-working-with-excel-and-python/


import os
from openpyxl import load_workbook


os.chdir('C:/Users/Sugar2/Documents/py/Python intro')

# load wb
wb = load_workbook('inFile.xlsx')
wb.save('outFile.xlsx')              # Saving. Use it whenever you want to see the file in the latest state
wb.sheetnames

ws = wb["Second"]             # Sheet by name
ws = wb[wb.sheetnames[0]]     # Sheet by index
ws.title                      # Sheet name
wb.create_sheet("April")      # New sheet with name
wb.create_sheet("January", 0) # New sheet with name and position
del wb["April"]               # Remove sheet by name
ws.sheet_properties.tabColor = "0072BA" # Tab color

lastRow = ws.max_row          # last row number
lastCol = ws.max_column       # last column number
len(ws['C'])                  # row number of column


a1 = ws.cell(1,1).value       # cell value
a2 = ws['A2'].value           # cell value

ws.cell(3,7).value = 500      # set cell value
ws['H31'].value = "=SUM(H1:H30)" # set cell value

ws.merge_cells('A1:B2')       # merge cells

# append rows to sheet
rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    ws.append(row)


# iterate rows and cols
    
for row in ws.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()   

for row in ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print() 


# Formatting
from openpyxl.styles import Alignment

cell = ws.cell(row=1, column=1)
cell.value = 'Sunny day'
cell.alignment = Alignment(horizontal='center', vertical='center')

cell = ws.cell(31, 8)
cell.font = cell.fonts.copy(bold=True)   
    

# Charting
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)


sheet = wb['January']

rows = [
    ("USA", 46),
    ("China", 38),
    ("UK", 29),
    ("Russia", 22),
    ("South Korea", 13),
    ("Germany", 11)
]

for row in rows:
    sheet.append(row)
    
    
    
data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
categs = Reference(sheet, min_col=1, min_row=1, max_row=6)

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categs)
chart.varyColors = True
chart.title = "Olympic Gold medals in London"
sheet.add_chart(chart, "A8")   



# xlrd. To read excel as in openpyxl but not to write. xlwt and Xlsxwriter, csv

import xlrd
wb = xlrd.open_workbook('outFile.xlsx')
sheet_names = wb.sheet_names()
wb.sheet_by_name(sheet_names[0])
ws = wb.sheet_by_index(0)
ws.row(0)
ws.ncols
ws.name
wb.nsheets
ws.row_values(0)
ws.cell(0,0)
ws.cell(0,0).value

ws.row_slice(rowx=0, start_colx=0, end_colx=2)










